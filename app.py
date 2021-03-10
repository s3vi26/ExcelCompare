import os
from flask import Flask, request
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
import logging

logging.basicConfig(level=logging.INFO)

logger = logging.getLogger('HELLO WORLD')

UPLOAD_FOLDER = '/Users/Sevi/code/ExcelCompare'
ALLOWED_EXTENSIONS = set(['xlsx'])

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

@app.route("/")
def hello():
  return "Hello World! This is our first endpoint"

@app.route("/upload", methods=['POST'])
def uploading():
  # 2 excel files
  if request.method == 'POST':
    print('is a POST method')
    print(request.files)
    if 'fileA' not in request.files or 'fileB' not in request.files:
      return 'files do not exist'
    
    target=os.path.join(UPLOAD_FOLDER,'backend_docs')
    
    if not os.path.isdir(target):
      os.mkdir(target)
    
    fileA = request.files['fileA']
    # To save file to a destination
    filename = secure_filename(fileA.filename)
    destination="/".join([target, filename])
    fileA.save(destination)
    workbookA = load_workbook(filename=f'backend_docs/{filename}')
    print(workbookA.sheetnames)
    sheetA = workbookA.active
    print(sheetA)
    my_listA = []
    for row in sheetA.iter_rows(min_row=1, min_col=1, values_only=True):
      filteredA = filter(None, row)
      for value in filteredA:
        my_listA.append(value)

    print(my_listA)
    
    fileB = request.files['fileB']
    # To save file to a destination
    filename = secure_filename(fileB.filename)
    destination="/".join([target, filename])
    fileB.save(destination)

    workbookB = load_workbook(filename=f'backend_docs/{filename}')
    print(workbookB.sheetnames)
    sheetB = workbookB.active
    print(sheetB)
    my_listB = []
    for row in sheetB.iter_rows(min_row=1, min_col=1, values_only=True):
      filteredB = filter(None, row)
      for value in filteredB:
        my_listB.append(value)

    print(my_listB)

    # write a fucntion to compare headers/comuns/rows
    # maybe think about if a DB is needed here? 
    
    return 'you hit the endpoint. files saved'
  else:
    return 'This method needs to be POST'

if __name__ == '__main__':
  # run app in debug mode on port 5000
  app.run(debug=True, port=5000)